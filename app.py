import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
from time import sleep

planilha_clientes = openpyxl.load_workbook('dados_clientes.xlsx')
pagina_clientes = planilha_clientes['Sheet1']

for linha in pagina_clientes.iter_rows(min_row=2,values_only=True):
   nome, valor, cpf, vencimento = linha
   print(nome)
   print(valor)
   print(cpf)
   print(vencimento)

driver = webdriver.Chrome()
driver.get('')
sleep(5)
campo_pesquisa = driver.find_element(By.XPATH,"")
sleep(1)
campo_pesquisa.send_keys(cpf)
sleep(1)

botao_pesquisar = driver.find_element(By.XPATH, "")
sleep(1)
botao_pesquisar.click()
sleep(5)
campo_pesquisa = driver.find_element(By.XPATH,"")
sleep(1)
campo_pesquisa.send_keys(cpf)
sleep(1)

botao_pesquisar = driver.find_element(By.XPATH, "")
sleep(1)
botao_pesquisar.click()

status = driver.find_element(By.XPATH, "//span[@id='statusLabel']")
if status.text == 'em dia':
   data_pagamento = driver.find_element(By.XPATH, "//p[@id='paymentDate']")
   metodo_pagamento = driver.find_element(By.XPATH, "//p[@id='paymentMethod']")
   
   pagina_fechamento.append([nome, valor, cpf, vencimento, 'em dia', 'xxx','xxx'])
   
else:
   planilha_fechamento = openpyxl.load_workbook(' planilha_fechamento.xlsx')
   pagina_fechamento = planilha_fechamento['Sheet1']
   
   pagina_fechamento.append([nome, valor, cpf, vencimento, 'pendente'])
   

